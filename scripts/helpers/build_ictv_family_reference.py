#!/usr/bin/env python3
"""Create the family-level genome display reference from an ICTV MSL workbook.

This utility deliberately preserves heterogeneous families as MIXED.  The
dashboard must never infer a single genome type from a partial contig or from
an incomplete set of species-level ICTV records.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import urlopen


CURRENT_MSL_URL = "https://ictv.global/sites/default/files/MSL/ICTV_Master_Species_List_2025_MSL41.v1.xlsx"
CURRENT_RELEASE = "ICTV MSL41.v1 (2025)"


def display_group(genome_values: set[str]) -> str:
    groups: set[str] = set()
    for value in genome_values:
        if value.endswith("-RT"):
            groups.add("RT")
        elif "DNA" in value:
            groups.add("DNA")
        elif "RNA" in value:
            groups.add("RNA")
        else:
            groups.add("UNCLASSIFIED")
    return next(iter(groups)) if len(groups) == 1 else "MIXED"


def detail(group: str, values: set[str]) -> str:
    raw = "; ".join(sorted(values))
    labels = {
        "DNA": "DNA virus",
        "RNA": "RNA virus",
        "RT": "Reverse-transcribing virus",
        "MIXED": "Family-level genome heterogeneity",
        "UNCLASSIFIED": "Unclassified",
    }
    return f"{labels[group]}; {raw}"


def download(url: str) -> Path:
    temporary = Path(tempfile.mkdtemp(prefix="ictv_msl_")) / "ictv_msl.xlsx"
    with urlopen(url, timeout=90) as response, temporary.open("wb") as handle:
        handle.write(response.read())
    return temporary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--input", type=Path, help="Downloaded ICTV MSL .xlsx file")
    source.add_argument("--download-current", action="store_true", help="Download the current bundled ICTV MSL URL")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--metadata", type=Path, required=True)
    parser.add_argument("--source-release", default=CURRENT_RELEASE)
    parser.add_argument("--source-url", default="https://ictv.global/msl")
    args = parser.parse_args()

    workbook_path = download(CURRENT_MSL_URL) if args.download_current else args.input
    if not workbook_path or not workbook_path.is_file():
        raise SystemExit("ICTV MSL workbook was not found")
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required. Run: pip install openpyxl==3.1.5") from exc

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "MSL" not in workbook.sheetnames:
        raise SystemExit("The workbook does not contain the ICTV 'MSL' sheet")
    sheet = workbook["MSL"]
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    try:
        family_index = header.index("Family")
        genome_index = header.index("Genome")
    except ValueError as exc:
        raise SystemExit("The MSL sheet is missing the required Family or Genome column") from exc

    values: dict[str, set[str]] = defaultdict(set)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        family = str(row[family_index] or "").strip()
        genome = str(row[genome_index] or "").strip()
        if family and genome:
            values[family].add(genome)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["family", "genome_group", "genome_detail", "ictv_genome_values", "dictionary_status", "source_release", "source_url"],
            delimiter="\t",
        )
        writer.writeheader()
        for family, genomes in sorted(values.items(), key=lambda item: item[0].lower()):
            group = display_group(genomes)
            writer.writerow({
                "family": family,
                "genome_group": group,
                "genome_detail": detail(group, genomes),
                "ictv_genome_values": "; ".join(sorted(genomes)),
                "dictionary_status": "review_required" if group == "MIXED" else "reference_mapped",
                "source_release": args.source_release,
                "source_url": args.source_url,
            })

    args.metadata.parent.mkdir(parents=True, exist_ok=True)
    args.metadata.write_text(json.dumps({
        "reference_name": "ICTV Master Species List family genome reference",
        "source_release": args.source_release,
        "source_url": args.source_url,
        "source_filename": workbook_path.name,
        "source_sha256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        "family_count": len(values),
        "generation_rule": "Group exact ICTV MSL Genome values by family; cross-group families are MIXED/review_required.",
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"[INFO] Wrote {len(values)} ICTV family rows: {args.output}")
    print(f"[INFO] Wrote reference metadata: {args.metadata}")


if __name__ == "__main__":
    main()
