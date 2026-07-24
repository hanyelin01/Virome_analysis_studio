#!/usr/bin/env python3
"""Build a version-pinned ICTV VMR protein reference from NCBI CDS records.

This administrator utility downloads protein translations only for accession
numbers listed in one frozen ICTV VMR spreadsheet.  Every retained protein is
linked to one VMR record in the emitted TSV.  Missing, ambiguous and no-CDS
accessions are retained in audit TSVs instead of being silently discarded.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen


ROOT = Path(__file__).resolve().parents[1]
EUTILS_EFETCH = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
ACCESSION = re.compile(r"\b(?:[A-Z]{1,6}_?\d{5,12})(?:\.\d+)?\b")
PROTEIN_ID = re.compile(r"\[protein_id=([^\]]+)\]")
NUCLEOTIDE_FROM_HEADER = re.compile(r"^lcl\|(.+?)_prot_")
SAFE_ID = re.compile(r"[^A-Za-z0-9_.|:-]+")


@dataclass(frozen=True)
class VmrRecord:
    vmr_id: str
    source_accession: str
    order: str
    family: str
    genus: str
    species: str
    genome: str
    baltimore_group: str
    isolate_role: str
    raw_accession_field: str


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_accession(value: str) -> str:
    value = value.strip().upper()
    return re.sub(r"\.\d+$", "", value)


def accessions_from_cell(value: object) -> list[str]:
    if value is None:
        return []
    seen: set[str] = set()
    result: list[str] = []
    for match in ACCESSION.findall(str(value).upper()):
        accession = normalize_accession(match)
        if accession not in seen:
            seen.add(accession)
            result.append(accession)
    return result


def baltimore_group(genome: str) -> str:
    normalized = genome.strip().replace(" ", "")
    if normalized == "dsDNA":
        return "I"
    if normalized.startswith("ssDNA"):
        return "II"
    if normalized == "dsRNA":
        return "III"
    if normalized == "ssRNA(+)":
        return "IV"
    if normalized in {"ssRNA(-)", "ssRNA(+/-)", "ssRNA(-);ssRNA(+/-)"}:
        return "V"
    if normalized == "ssRNA-RT":
        return "VI"
    if normalized == "dsDNA-RT":
        return "VII"
    return "unclassified"


def nonempty(value: object) -> str:
    text = str(value or "").strip()
    return text if text else "unclassified"


def load_vmr(path: Path) -> tuple[list[VmrRecord], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("openpyxl is required; run this builder with the project's .venv Python.") from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = [name for name in workbook.sheetnames if name.startswith("VMR ")]
    if len(sheets) != 1:
        raise SystemExit(f"Expected exactly one VMR data sheet, found: {', '.join(sheets) or 'none'}")
    sheet = workbook[sheets[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    required = {"Isolate ID", "Order", "Family", "Genus", "Species", "Exemplar or additional isolate", "Virus GENBANK accession", "Genome"}
    missing = sorted(required - set(headers))
    if missing:
        raise SystemExit(f"VMR spreadsheet is missing expected columns: {', '.join(missing)}")
    positions = {header: headers.index(header) for header in required}
    records: list[VmrRecord] = []
    ignored: list[dict[str, str]] = []
    for number, row in enumerate(rows, start=2):
        raw_accession = row[positions["Virus GENBANK accession"]]
        ids = accessions_from_cell(raw_accession)
        if not ids:
            ignored.append({"row_number": str(number), "vmr_id": nonempty(row[positions["Isolate ID"]]), "reason": "no_parseable_genbank_accession", "raw_accession_field": str(raw_accession or "")})
            continue
        row_values = {
            "vmr_id": nonempty(row[positions["Isolate ID"]]),
            "order": nonempty(row[positions["Order"]]),
            "family": nonempty(row[positions["Family"]]),
            "genus": nonempty(row[positions["Genus"]]),
            "species": nonempty(row[positions["Species"]]),
            "genome": nonempty(row[positions["Genome"]]),
            "isolate_role": nonempty(row[positions["Exemplar or additional isolate"]]),
            "raw_accession_field": str(raw_accession or ""),
        }
        for accession in ids:
            records.append(VmrRecord(source_accession=accession, baltimore_group=baltimore_group(row_values["genome"]), **row_values))
    if not records:
        raise SystemExit("No parseable GenBank accession was found in the VMR spreadsheet")
    return records, ignored


def taxonomy_key(record: VmrRecord) -> tuple[str, str, str, str, str, str]:
    return record.order, record.family, record.genus, record.species, record.genome, record.baltimore_group


def canonical_records(records: list[VmrRecord]) -> tuple[dict[str, VmrRecord], list[dict[str, str]]]:
    grouped: dict[str, list[VmrRecord]] = defaultdict(list)
    for record in records:
        grouped[record.source_accession].append(record)
    accepted: dict[str, VmrRecord] = {}
    excluded: list[dict[str, str]] = []
    for accession, items in sorted(grouped.items()):
        keys = {taxonomy_key(item) for item in items}
        if len(keys) != 1:
            excluded.append({
                "source_accession": accession,
                "reason": "conflicting_vmr_taxonomy",
                "vmr_ids": ";".join(sorted({item.vmr_id for item in items})),
                "species": ";".join(sorted({item.species for item in items})),
            })
            continue
        accepted[accession] = sorted(items, key=lambda item: (item.isolate_role != "E", item.vmr_id))[0]
    return accepted, excluded


def chunks(values: list[str], size: int) -> Iterable[list[str]]:
    for index in range(0, len(values), size):
        yield values[index:index + size]


def fetch_batch(accessions: list[str], *, email: str, delay: float) -> str:
    parameters = {"db": "nuccore", "id": ",".join(accessions), "rettype": "fasta_cds_aa", "retmode": "text", "tool": "virome_analysis_studio"}
    if email:
        parameters["email"] = email
    request = Request(f"{EUTILS_EFETCH}?{urlencode(parameters)}", headers={"User-Agent": "ViromeAnalysisStudio/2.0 (ICTV VMR reference builder)"})
    for attempt in range(5):
        try:
            with urlopen(request, timeout=180) as response:
                payload = response.read().decode("utf-8", errors="replace")
            if payload.lstrip().startswith("<?xml") or "<ERROR>" in payload:
                raise RuntimeError(payload[:500])
            time.sleep(delay)
            return payload
        except (HTTPError, URLError, TimeoutError, RuntimeError) as error:
            if attempt == 4:
                raise RuntimeError(f"NCBI EFetch failed for batch beginning {accessions[0]}: {error}") from error
            time.sleep(max(delay, 1.0) * (2 ** attempt))
    raise AssertionError("unreachable")


def fasta_records(text: str) -> Iterable[tuple[str, str]]:
    header: str | None = None
    sequence: list[str] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                yield header, "".join(sequence)
            header, sequence = line[1:], []
        else:
            sequence.append(line)
    if header is not None:
        yield header, "".join(sequence)


def accession_from_header(header: str) -> str | None:
    token = header.split(maxsplit=1)[0]
    match = NUCLEOTIDE_FROM_HEADER.match(token)
    if match:
        return normalize_accession(match.group(1))
    match = ACCESSION.search(header.upper())
    return normalize_accession(match.group(0)) if match else None


def protein_from_header(header: str, ordinal: int) -> str:
    match = PROTEIN_ID.search(header)
    if match:
        return SAFE_ID.sub("_", match.group(1).strip())
    return f"cds_{ordinal:06d}"


def write_tsv(path: Path, fields: list[str], rows: Iterable[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build an audited ICTV VMR protein reference staging set.")
    parser.add_argument("--vmr-xlsx", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True, help="New or resumable VMR build directory")
    parser.add_argument("--version", required=True, help="Pinned ICTV VMR release, e.g. VMR_MSL41.v1.20260721")
    parser.add_argument("--email", default="", help="Optional contact email sent to NCBI E-utilities")
    parser.add_argument("--batch-size", type=int, default=50)
    parser.add_argument("--request-delay", type=float, default=0.34, help="Seconds between successful NCBI requests without an API key")
    parser.add_argument("--resume", action="store_true", help="Reuse successfully downloaded batch FASTA files")
    parser.add_argument("--max-accessions", type=int, help="Testing only: cap the deterministic accession list")
    args = parser.parse_args()
    if not args.vmr_xlsx.is_file():
        raise SystemExit(f"VMR spreadsheet is unavailable: {args.vmr_xlsx}")
    if args.batch_size < 1 or args.batch_size > 200 or args.request_delay < 0.34:
        raise SystemExit("batch-size must be 1–200 and request-delay must be at least 0.34 seconds")
    if args.output_dir.exists() and any(args.output_dir.iterdir()) and not args.resume:
        raise SystemExit(f"Output directory is not empty; use --resume after verifying it: {args.output_dir}")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    retrieval = args.output_dir / "retrieval_batches"
    retrieval.mkdir(exist_ok=True)
    records, no_accession = load_vmr(args.vmr_xlsx)
    accepted, conflicts = canonical_records(records)
    accession_list = sorted(accepted)
    if args.max_accessions is not None:
        accession_list = accession_list[:args.max_accessions]
    (args.output_dir / "source_accessions.txt").write_text("\n".join(accession_list) + "\n", encoding="utf-8")
    source_copy = args.output_dir / args.vmr_xlsx.name
    if not source_copy.exists():
        shutil.copy2(args.vmr_xlsx, source_copy)
    batches = list(chunks(accession_list, args.batch_size))
    for index, batch in enumerate(batches, start=1):
        target = retrieval / f"batch_{index:05d}.faa"
        if target.is_file() and target.stat().st_size and args.resume:
            continue
        payload = fetch_batch(batch, email=args.email, delay=args.request_delay)
        temporary = target.with_suffix(".faa.partial")
        temporary.write_text(payload, encoding="utf-8")
        temporary.replace(target)
        print(f"[INFO] Downloaded NCBI batch {index}/{len(batches)} ({len(batch)} accession(s))", flush=True)

    metadata_path = args.output_dir / "ictv_reviewed_metadata.tsv"
    fasta_path = args.output_dir / "ictv_proteins.faa"
    unresolved: list[dict[str, str]] = list(no_accession) + list(conflicts)
    seen_accessions: set[str] = set()
    seen_reference_ids: set[str] = set()
    protein_count = 0
    metadata_fields = ["reference_id", "family", "genus", "species", "baltimore_group", "order", "genome_composition", "vmr_id", "source_accession", "protein_accession", "isolate_role", "vmr_version"]
    with metadata_path.open("w", encoding="utf-8", newline="") as metadata_handle, fasta_path.open("w", encoding="utf-8") as fasta_handle:
        metadata_writer = csv.DictWriter(metadata_handle, fieldnames=metadata_fields, delimiter="\t")
        metadata_writer.writeheader()
        for batch_path in sorted(retrieval.glob("batch_*.faa")):
            for ordinal, (header, sequence) in enumerate(fasta_records(batch_path.read_text(encoding="utf-8", errors="replace")), start=1):
                source_accession = accession_from_header(header)
                if source_accession not in accepted or source_accession not in accession_list:
                    continue
                seen_accessions.add(source_accession)
                record = accepted[source_accession]
                protein = protein_from_header(header, ordinal)
                reference_id = SAFE_ID.sub("_", f"{record.vmr_id}|{source_accession}|{protein}")
                if reference_id in seen_reference_ids or not sequence:
                    continue
                seen_reference_ids.add(reference_id)
                fasta_handle.write(f">{reference_id}\n")
                for offset in range(0, len(sequence), 80):
                    fasta_handle.write(sequence[offset:offset + 80] + "\n")
                metadata_writer.writerow({
                    "reference_id": reference_id,
                    "family": record.family,
                    "genus": record.genus,
                    "species": record.species,
                    "baltimore_group": record.baltimore_group,
                    "order": record.order,
                    "genome_composition": record.genome,
                    "vmr_id": record.vmr_id,
                    "source_accession": source_accession,
                    "protein_accession": protein,
                    "isolate_role": record.isolate_role,
                    "vmr_version": args.version,
                })
                protein_count += 1
    for accession in accession_list:
        if accession not in seen_accessions:
            unresolved.append({"source_accession": accession, "reason": "no_cds_protein_returned_by_ncbi", "vmr_id": accepted[accession].vmr_id, "species": accepted[accession].species})
    unresolved_path = args.output_dir / "ictv_unresolved_accessions.tsv"
    fields = sorted({key for row in unresolved for key in row}) or ["source_accession", "reason"]
    write_tsv(unresolved_path, fields, unresolved)
    manifest = {
        "schema_version": 1,
        "ictv_vmr_version": args.version,
        "vmr_source_file": source_copy.name,
        "vmr_source_sha256": sha256(source_copy),
        "requested_accession_count": len(accession_list),
        "source_record_count": len(records),
        "conflicting_accession_count": len(conflicts),
        "accession_with_cds_count": len(seen_accessions),
        "protein_record_count": protein_count,
        "unresolved_record_count": len(unresolved),
        "metadata_file": metadata_path.name,
        "metadata_sha256": sha256(metadata_path),
        "protein_fasta": fasta_path.name,
        "protein_fasta_sha256": sha256(fasta_path),
    }
    (args.output_dir / "ictv_vmr_retrieval_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if protein_count == 0:
        raise SystemExit("No protein CDS records were retrieved; inspect retrieval batches and unresolved TSV")
    print(f"[INFO] Staged {protein_count} ICTV VMR protein record(s) from {len(seen_accessions)}/{len(accession_list)} accession(s): {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
